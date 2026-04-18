"""
relatorios/services.py
======================
CAMADA DE SERVIÇO (Service Layer)

Por que criar um services.py separado das views?
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Views devem ser "burras" — só receber requisição e devolver resposta
• A lógica complexa de negócio fica aqui, isolada e reutilizável
• Posso chamar este serviço de uma view HTML, de uma API, de um
  management command, de um teste — sem duplicar código

Este arquivo gera relatórios Excel com múltiplas abas usando
pandas + openpyxl. É o equivalente profissional do Ctrl+C / Ctrl+V
de dados para uma planilha manual.
"""

import io
from datetime import date, timedelta, datetime
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    numbers as xl_numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone


# ═══════════════════════════════════════════════════════════════
# CONSTANTES DE ESTILO — Define a identidade visual do Excel
# ═══════════════════════════════════════════════════════════════
COR_PRIMARIA    = "E8520A"   # laranja LanchoPro
COR_HEADER      = "1A1A2E"   # azul escuro
COR_SUBHEADER   = "2D2D4E"
COR_BRANCO      = "FFFFFF"
COR_CINZA_CLARO = "F5F5FA"
COR_CINZA_MED   = "E0E0EE"
COR_VERDE       = "0D6B39"
COR_VERMELHO    = "C0392B"
COR_AMARELO     = "F39C12"

FILL_HEADER    = PatternFill("solid", fgColor=COR_HEADER)
FILL_SUBHEADER = PatternFill("solid", fgColor=COR_PRIMARIA)
FILL_CINZA     = PatternFill("solid", fgColor=COR_CINZA_CLARO)
FILL_VERDE     = PatternFill("solid", fgColor="E8F8F0")
FILL_VERMELHO  = PatternFill("solid", fgColor="FDECEA")

FONT_TITULO    = Font(name="Arial", size=16, bold=True, color=COR_HEADER)
FONT_HEADER    = Font(name="Arial", size=10, bold=True, color=COR_BRANCO)
FONT_SUBHEADER = Font(name="Arial", size=10, bold=True, color=COR_BRANCO)
FONT_NORMAL    = Font(name="Arial", size=10)
FONT_BOLD      = Font(name="Arial", size=10, bold=True)
FONT_VERDE     = Font(name="Arial", size=10, bold=True, color=COR_VERDE)
FONT_VERMELHO  = Font(name="Arial", size=10, bold=True, color=COR_VERMELHO)
FONT_LARANJA   = Font(name="Arial", size=10, bold=True, color=COR_PRIMARIA)

BORDA_FINA = Border(
    left=Side(style='thin', color="CCCCDD"),
    right=Side(style='thin', color="CCCCDD"),
    top=Side(style='thin', color="CCCCDD"),
    bottom=Side(style='thin', color="CCCCDD"),
)

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal='right',  vertical='center')

FMT_MOEDA   = 'R$ #,##0.00'
FMT_PERCENT = '0.00"%"'
FMT_INT     = '#,##0'
FMT_DATA    = 'DD/MM/YYYY'
FMT_DATAHORA= 'DD/MM/YYYY HH:MM'


# ═══════════════════════════════════════════════════════════════
# FUNÇÕES UTILITÁRIAS
# ═══════════════════════════════════════════════════════════════

def estilizar_celula(cell, font=None, fill=None, alignment=None,
                     border=None, number_format=None):
    """Aplica estilos a uma célula de forma conveniente."""
    if font:       cell.font = font
    if fill:       cell.fill = fill
    if alignment:  cell.alignment = alignment
    if border:     cell.border = border
    if number_format: cell.number_format = number_format


def autofit_colunas(ws, min_width=8, max_width=50):
    """Ajusta largura das colunas automaticamente pelo conteúdo."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def linha_header(ws, row, colunas, altura=30):
    """Cria linha de cabeçalho estilizada."""
    for col_idx, (texto, largura) in enumerate(colunas, start=1):
        cell = ws.cell(row=row, column=col_idx, value=texto)
        estilizar_celula(cell,
            font=FONT_HEADER,
            fill=FILL_HEADER,
            alignment=ALIGN_CENTER,
            border=BORDA_FINA,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = largura
    ws.row_dimensions[row].height = altura


def linha_subheader(ws, row, colunas):
    """Linha de sub-cabeçalho (laranja)."""
    for col_idx, texto in enumerate(colunas, start=1):
        cell = ws.cell(row=row, column=col_idx, value=texto)
        estilizar_celula(cell,
            font=FONT_SUBHEADER,
            fill=FILL_SUBHEADER,
            alignment=ALIGN_CENTER,
            border=BORDA_FINA,
        )


def linha_dados(ws, row, valores, is_par=False):
    """Linha de dados com zebra striping."""
    fill = FILL_CINZA if is_par else None
    for col_idx, valor in enumerate(valores, start=1):
        cell = ws.cell(row=row, column=col_idx, value=valor)
        estilizar_celula(cell,
            font=FONT_NORMAL,
            fill=fill,
            alignment=ALIGN_LEFT,
            border=BORDA_FINA,
        )
    return ws.row_dimensions[row]


def titulo_aba(ws, titulo, subtitulo, n_colunas, logo="🍔"):
    """Cria cabeçalho visual da aba com título e data de geração."""
    # Linha 1: Logo + título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_colunas)
    cell = ws.cell(row=1, column=1, value=f"{logo}  LanchoPro — {titulo}")
    estilizar_celula(cell, font=FONT_TITULO, alignment=ALIGN_LEFT)
    ws.row_dimensions[1].height = 35

    # Linha 2: Subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_colunas)
    cell = ws.cell(row=2, column=1, value=subtitulo)
    estilizar_celula(cell,
        font=Font(name="Arial", size=10, italic=True, color="666688"),
        alignment=ALIGN_LEFT,
    )

    # Linha 3: Data/hora da geração
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_colunas)
    agora = datetime.now().strftime("%d/%m/%Y às %H:%M")
    cell = ws.cell(row=3, column=1, value=f"Gerado em: {agora}")
    estilizar_celula(cell,
        font=Font(name="Arial", size=9, italic=True, color="999999"),
        alignment=ALIGN_LEFT,
    )

    # Linha 4: espaço
    ws.row_dimensions[4].height = 8
    return 5   # próxima linha disponível


# ═══════════════════════════════════════════════════════════════
# GERADOR PRINCIPAL — RELATÓRIO COMPLETO
# ═══════════════════════════════════════════════════════════════

def gerar_relatorio_excel(data_inicio=None, data_fim=None):
    """
    Gera um workbook Excel completo com 5 abas:
    1. Resumo Executivo
    2. Vendas Detalhadas
    3. Produtos Mais Vendidos
    4. Estoque Atual
    5. Contas em Aberto (Fiado)

    Retorna: BytesIO (buffer em memória, pronto para download)
    """
    # Importações tardias (evita circular imports)
    from vendas.models import Venda, ItemVenda
    from estoque.models import Produto, Categoria
    from clientes.models import Cliente, ContaAberta
    from caixa.models import Caixa

    # ── Período ─────────────────────────────────────────────
    hoje = date.today()
    if not data_inicio:
        data_inicio = hoje.replace(day=1)   # início do mês
    if not data_fim:
        data_fim = hoje

    periodo_str = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"

    # ── Dados base ───────────────────────────────────────────
    vendas_qs = Venda.objects.filter(
        criado_em__date__gte=data_inicio,
        criado_em__date__lte=data_fim,
        status='CONCLUIDA',
    ).select_related('operador').prefetch_related('itens')

    wb = Workbook()
    wb.remove(wb.active)   # remove aba padrão em branco

    # ════════════════════════════════════════════════════════
    # ABA 1 — RESUMO EXECUTIVO
    # ════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("📊 Resumo Executivo")
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A6"

    row = titulo_aba(ws1, "Resumo Executivo", f"Período: {periodo_str}", 4)

    # Calcular métricas
    total_vendas   = vendas_qs.aggregate(t=Sum('total'))['t'] or Decimal('0')
    qtd_vendas     = vendas_qs.count()
    ticket_medio   = total_vendas / qtd_vendas if qtd_vendas else Decimal('0')
    total_fiado_v  = vendas_qs.filter(tipo_pagamento='FIADO').aggregate(t=Sum('total'))['t'] or Decimal('0')

    total_fiado_ab = ContaAberta.objects.filter(pago=False).aggregate(t=Sum('total'))['t'] or Decimal('0')
    total_estoque  = Produto.objects.filter(ativo=True).aggregate(
        t=Sum('quantidade')
    )['t'] or 0
    valor_estoque  = sum(
        (p.quantidade * p.preco)
        for p in Produto.objects.filter(ativo=True)
    )
    prod_baixo     = sum(1 for p in Produto.objects.filter(ativo=True) if p.estoque_baixo)

    # Cards de métricas (tabela 2x2)
    metricas = [
        ("💰  Total de Vendas no Período", float(total_vendas), FMT_MOEDA, FONT_VERDE),
        ("🛒  Número de Transações",        qtd_vendas,          FMT_INT,   FONT_BOLD),
        ("🎯  Ticket Médio",                float(ticket_medio),  FMT_MOEDA, FONT_LARANJA),
        ("📋  Vendas Fiado (no período)",   float(total_fiado_v), FMT_MOEDA, FONT_BOLD),
        ("⚠️  Fiado Total em Aberto",        float(total_fiado_ab),FMT_MOEDA, FONT_VERMELHO),
        ("📦  Produtos em Estoque",          total_estoque,        FMT_INT,   FONT_BOLD),
        ("💎  Valor do Estoque",             float(valor_estoque), FMT_MOEDA, FONT_VERDE),
        ("🔴  Produtos com Estoque Baixo",   prod_baixo,           FMT_INT,   FONT_VERMELHO),
    ]

    # Cabeçalhos da tabela de métricas
    row += 1
    for c, txt in [(1, "Métrica"), (3, "Valor")]:
        ws1.merge_cells(start_row=row, start_column=c, end_row=row, end_column=c+1)
        cell = ws1.cell(row=row, column=c, value=txt)
        estilizar_celula(cell, font=FONT_HEADER, fill=FILL_SUBHEADER,
                         alignment=ALIGN_CENTER, border=BORDA_FINA)
    row += 1

    for i, (label, valor, fmt, font) in enumerate(metricas):
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c1 = ws1.cell(row=row, column=1, value=label)
        fill = FILL_CINZA if i % 2 == 0 else None
        estilizar_celula(c1, font=FONT_NORMAL, fill=fill,
                         alignment=ALIGN_LEFT, border=BORDA_FINA)

        ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c2 = ws1.cell(row=row, column=3, value=valor)
        estilizar_celula(c2, font=font, fill=fill,
                         alignment=ALIGN_RIGHT, border=BORDA_FINA,
                         number_format=fmt)
        ws1.row_dimensions[row].height = 22
        row += 1

    # Gráfico — Vendas por forma de pagamento
    row += 2
    ws1.cell(row=row, column=1, value="Vendas por Forma de Pagamento").font = FONT_BOLD
    row += 1

    pgtos = vendas_qs.values('tipo_pagamento').annotate(
        total=Sum('total'), qtd=Count('numero')
    ).order_by('-total')

    pgto_labels = {
        'DINHEIRO': 'Dinheiro', 'PIX': 'PIX',
        'CARTAO_DEBITO': 'Débito', 'CARTAO_CREDITO': 'Crédito', 'FIADO': 'Fiado',
    }
    header_pgto = [("Forma de Pagamento", 25), ("Quantidade", 14), ("Total (R$)", 18), ("% do Total", 14)]
    linha_header(ws1, row, header_pgto)
    row += 1

    total_float = float(total_vendas) or 1
    for i, p in enumerate(pgtos):
        pct = (float(p['total']) / total_float) * 100
        linha_dados(ws1, row,
            [pgto_labels.get(p['tipo_pagamento'], p['tipo_pagamento']),
             p['qtd'], float(p['total']), pct],
            is_par=i % 2 == 0
        )
        ws1.cell(row=row, column=3).number_format = FMT_MOEDA
        ws1.cell(row=row, column=4).number_format = FMT_PERCENT
        row += 1

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 15

    # ════════════════════════════════════════════════════════
    # ABA 2 — VENDAS DETALHADAS
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("🧾 Vendas Detalhadas")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A6"

    row = titulo_aba(ws2, "Vendas Detalhadas", f"Período: {periodo_str}", 8)

    cols = [
        ("#",              6),  ("Data/Hora",       18), ("Cliente",         20),
        ("Pagamento",      14), ("Itens",            8),  ("Subtotal",        14),
        ("Desconto",       12), ("Total",            14),
    ]
    linha_header(ws2, row, cols)
    row += 1

    for i, venda in enumerate(vendas_qs.order_by('-criado_em')):
        valores = [
            venda.numero,
            venda.criado_em.strftime("%d/%m/%Y %H:%M"),
            venda.cliente_nome or "—",
            pgto_labels.get(venda.tipo_pagamento, venda.tipo_pagamento),
            venda.itens.count(),
            float(venda.subtotal),
            float(venda.desconto),
            float(venda.total),
        ]
        linha_dados(ws2, row, valores, is_par=i % 2 == 0)

        for col_idx, fmt in [(6, FMT_MOEDA), (7, FMT_MOEDA), (8, FMT_MOEDA)]:
            ws2.cell(row=row, column=col_idx).number_format = fmt
        ws2.cell(row=row, column=8).font = FONT_BOLD
        row += 1

    # Linha de totais
    row += 1
    ws2.cell(row=row, column=1, value="TOTAL GERAL").font = FONT_BOLD
    total_cell = ws2.cell(row=row, column=8, value=float(total_vendas))
    estilizar_celula(total_cell, font=FONT_VERDE, number_format=FMT_MOEDA,
                     fill=PatternFill("solid", fgColor="E8F8F0"), border=BORDA_FINA)

    # ════════════════════════════════════════════════════════
    # ABA 3 — PRODUTOS MAIS VENDIDOS
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("🏆 Mais Vendidos")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A6"

    row = titulo_aba(ws3, "Produtos Mais Vendidos", f"Período: {periodo_str}", 5)

    top_produtos = ItemVenda.objects.filter(
        venda__criado_em__date__gte=data_inicio,
        venda__criado_em__date__lte=data_fim,
        venda__status='CONCLUIDA',
    ).values('produto_nome').annotate(
        qtd_total=Sum('quantidade'),
        receita=Sum('total'),
        n_vendas=Count('venda', distinct=True),
    ).order_by('-qtd_total')

    cols3 = [
        ("Ranking",        10), ("Produto",          30), ("Qtd Vendida",     14),
        ("Nº de Vendas",   14), ("Receita Total",    16),
    ]
    linha_header(ws3, row, cols3)
    row += 1

    receita_total = sum(float(p['receita']) for p in top_produtos) or 1
    for i, p in enumerate(top_produtos):
        ranking = i + 1
        medal = "🥇" if ranking == 1 else "🥈" if ranking == 2 else "🥉" if ranking == 3 else str(ranking)
        linha_dados(ws3, row,
            [medal, p['produto_nome'], p['qtd_total'], p['n_vendas'], float(p['receita'])],
            is_par=i % 2 == 0
        )
        ws3.cell(row=row, column=5).number_format = FMT_MOEDA
        if ranking <= 3:
            for col in range(1, 6):
                ws3.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF8E0")
        row += 1

    # Gráfico de barras
    if top_produtos:
        chart = BarChart()
        chart.type = "col"
        chart.title = "Top 10 Produtos por Receita"
        chart.y_axis.title = "Receita (R$)"
        chart.x_axis.title = "Produto"
        chart.style = 10
        chart.height = 12
        chart.width = 20

        data_ref = Reference(ws3, min_col=5, min_row=row-min(len(list(top_produtos)), 10)-1,
                              max_row=row-1)
        cats_ref = Reference(ws3, min_col=2, min_row=row-min(len(list(top_produtos)), 10),
                              max_row=row-1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws3.add_chart(chart, f"A{row+3}")

    # ════════════════════════════════════════════════════════
    # ABA 4 — ESTOQUE ATUAL
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📦 Estoque Atual")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = "A6"

    row = titulo_aba(ws4, "Estoque Atual", f"Gerado em: {hoje.strftime('%d/%m/%Y')}", 7)

    cols4 = [
        ("Produto",        28), ("Categoria",        18), ("Qtd Estoque",    14),
        ("Estoque Mín.",   14), ("Status",           14), ("Preço Venda",    14),
        ("Valor em Estoque", 18),
    ]
    linha_header(ws4, row, cols4)
    row += 1

    produtos = Produto.objects.filter(ativo=True).select_related('categoria').order_by('categoria__nome', 'nome')
    for i, p in enumerate(produtos):
        em_falta = p.estoque_baixo
        status = "🔴 BAIXO" if em_falta else "🟢 OK"
        valor_estoque_produto = float(p.preco) * p.quantidade

        linha_dados(ws4, row,
            [p.nome, p.categoria.nome if p.categoria else "—",
             p.quantidade, p.estoque_minimo, status,
             float(p.preco), valor_estoque_produto],
            is_par=i % 2 == 0
        )
        ws4.cell(row=row, column=6).number_format = FMT_MOEDA
        ws4.cell(row=row, column=7).number_format = FMT_MOEDA

        if em_falta:
            for col in range(1, 8):
                c = ws4.cell(row=row, column=col)
                c.fill = FILL_VERMELHO
                c.font = Font(name="Arial", size=10, color=COR_VERMELHO)

        row += 1

    # Linha total de valor
    row += 1
    ws4.cell(row=row, column=6, value="VALOR TOTAL:").font = FONT_BOLD
    tot = ws4.cell(row=row, column=7, value=float(valor_estoque))
    estilizar_celula(tot, font=FONT_VERDE, number_format=FMT_MOEDA,
                     fill=PatternFill("solid", fgColor="E8F8F0"), border=BORDA_FINA)

    # ════════════════════════════════════════════════════════
    # ABA 5 — CONTAS EM ABERTO (FIADO)
    # ════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("💳 Fiado em Aberto")
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = "A6"

    row = titulo_aba(ws5, "Contas em Aberto (Fiado)", f"Gerado em: {hoje.strftime('%d/%m/%Y')}", 7)

    cols5 = [
        ("Cliente",        22), ("Telefone",         16), ("Descrição",       30),
        ("Total",          14), ("Já Pago",           14), ("Saldo Devedor",   16),
        ("Vencimento",     14),
    ]
    linha_header(ws5, row, cols5)
    row += 1

    contas = ContaAberta.objects.filter(pago=False).select_related('cliente').order_by('cliente__nome', 'criado_em')
    total_fiado_total = Decimal('0')
    for i, c in enumerate(contas):
        venc = c.data_vencimento.strftime("%d/%m/%Y") if c.data_vencimento else "—"
        vencido = c.data_vencimento and c.data_vencimento < hoje

        linha_dados(ws5, row,
            [c.cliente.nome, c.cliente.telefone or "—",
             c.descricao[:50], float(c.total), float(c.valor_pago),
             float(c.saldo_devedor), venc],
            is_par=i % 2 == 0
        )
        for col in [4, 5, 6]:
            ws5.cell(row=row, column=col).number_format = FMT_MOEDA

        if vencido:
            for col in range(1, 8):
                ws5.cell(row=row, column=col).fill = FILL_VERMELHO
        total_fiado_total += c.saldo_devedor
        row += 1

    row += 1
    ws5.cell(row=row, column=5, value="TOTAL DEVIDO:").font = FONT_BOLD
    tot5 = ws5.cell(row=row, column=6, value=float(total_fiado_total))
    estilizar_celula(tot5, font=FONT_VERMELHO, number_format=FMT_MOEDA,
                     fill=FILL_VERMELHO, border=BORDA_FINA)

    # ════════════════════════════════════════════════════════
    # ABA 6 — VENDAS POR DIA (SÉRIE TEMPORAL)
    # ════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("📈 Evolução Diária")
    ws6.sheet_view.showGridLines = False

    row = titulo_aba(ws6, "Evolução de Vendas por Dia", f"Período: {periodo_str}", 5)

    cols6 = [("Data", 14), ("Nº Vendas", 12), ("Total (R$)", 16), ("Ticket Médio", 16), ("Maior Venda", 14)]
    linha_header(ws6, row, cols6)
    row_chart_start = row + 1
    row += 1

    dias = []
    delta = data_fim - data_inicio
    for i in range(delta.days + 1):
        dia = data_inicio + timedelta(days=i)
        vendas_dia = vendas_qs.filter(criado_em__date=dia)
        agg = vendas_dia.aggregate(
            soma=Sum('total'), qtd=Count('numero')
        )
        total_dia = float(agg['soma'] or 0)
        qtd_dia   = agg['qtd'] or 0
        ticket    = total_dia / qtd_dia if qtd_dia else 0

        linha_dados(ws6, row,
            [dia.strftime("%d/%m/%Y"), qtd_dia, total_dia, ticket, total_dia],
            is_par=i % 2 == 0
        )
        ws6.cell(row=row, column=3).number_format = FMT_MOEDA
        ws6.cell(row=row, column=4).number_format = FMT_MOEDA
        ws6.cell(row=row, column=5).number_format = FMT_MOEDA
        dias.append(dia)
        row += 1

    # Gráfico de linha — evolução diária
    if len(dias) > 1:
        chart = LineChart()
        chart.title = "Evolução de Vendas Diárias"
        chart.y_axis.title = "Total (R$)"
        chart.x_axis.title = "Data"
        chart.style = 10
        chart.height = 14
        chart.width = 24

        data_ref = Reference(ws6, min_col=3, min_row=row_chart_start,
                              max_row=row_chart_start + len(dias) - 1)
        cats_ref = Reference(ws6, min_col=1, min_row=row_chart_start + 1,
                              max_row=row_chart_start + len(dias) - 1)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        chart.series[0].title.value = "Vendas Diárias"
        ws6.add_chart(chart, f"A{row+3}")

    # ═══════════════════════════════════════════
    # Propriedades do workbook
    # ═══════════════════════════════════════════
    wb.properties.title = f"LanchoPro — Relatório {periodo_str}"
    wb.properties.creator = "LanchoPro Sistema"
    wb.properties.description = "Relatório gerado automaticamente pelo sistema LanchoPro"

    # Salva em memória (BytesIO) — não cria arquivo no disco
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_relatorio_caixa_excel(caixa):
    """
    Gera relatório de um caixa específico com todas as movimentações.
    Retorna BytesIO pronto para download.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Caixa"
    ws.sheet_view.showGridLines = False

    data_str = caixa.data_abertura.strftime("%d/%m/%Y")
    row = titulo_aba(ws, f"Relatório de Caixa — {data_str}",
                     f"Operador: {caixa.operador.username if caixa.operador else '—'}", 5)

    # Resumo
    resumo = [
        ("Data de Abertura",   caixa.data_abertura.strftime("%d/%m/%Y %H:%M")),
        ("Data de Fechamento", caixa.data_fechamento.strftime("%d/%m/%Y %H:%M") if caixa.data_fechamento else "Em aberto"),
        ("Operador",           caixa.operador.username if caixa.operador else "—"),
        ("Valor Inicial",      float(caixa.valor_inicial)),
        ("Total de Vendas",    float(caixa.valor_vendas)),
        ("Total de Sangrias",  float(caixa.valor_sangria)),
        ("Total de Reforços",  float(caixa.valor_reforco)),
        ("Valor Final Contado",float(caixa.valor_final)),
        ("Saldo Esperado",     float(caixa.saldo_esperado)),
        ("Diferença",          float(caixa.valor_final) - float(caixa.saldo_esperado)),
    ]

    linha_header(ws, row, [("Campo", 28), ("Valor", 24)])
    row += 1
    for i, (campo, valor) in enumerate(resumo):
        cell_k = ws.cell(row=row, column=1, value=campo)
        cell_v = ws.cell(row=row, column=2, value=valor)
        fill = FILL_CINZA if i % 2 == 0 else None
        estilizar_celula(cell_k, font=FONT_BOLD, fill=fill, alignment=ALIGN_LEFT, border=BORDA_FINA)
        estilizar_celula(cell_v, font=FONT_NORMAL, fill=fill, alignment=ALIGN_RIGHT, border=BORDA_FINA)
        if isinstance(valor, float):
            cell_v.number_format = FMT_MOEDA
        row += 1

    # Movimentações
    row += 2
    ws.cell(row=row, column=1, value="Movimentações Detalhadas").font = FONT_BOLD
    row += 1
    cols_mov = [("Data/Hora", 20), ("Tipo", 18), ("Valor", 16), ("Operador", 18), ("Descrição", 30)]
    linha_header(ws, row, cols_mov)
    row += 1

    for i, mov in enumerate(caixa.movimentacoes.order_by('data_hora')):
        linha_dados(ws, row,
            [mov.data_hora.strftime("%d/%m/%Y %H:%M"),
             mov.get_tipo_display(),
             float(mov.valor),
             mov.operador.username if mov.operador else "—",
             mov.descricao or "—"],
            is_par=i % 2 == 0
        )
        ws.cell(row=row, column=3).number_format = FMT_MOEDA
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
